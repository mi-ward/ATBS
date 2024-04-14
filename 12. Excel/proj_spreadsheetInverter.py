
import openpyxl, sys

wb = openpyxl.load_workbook(sys.argv[1])
new_wb = openpyxl.Workbook()

sheet = wb.active
new_sheet = new_wb.active

for x in range(1, sheet.max_column+1):
    for y in range(1, sheet.max_row+1):
        new_sheet.cell(row=x,column=y).value = sheet.cell(row=y,column=x).value

new_wb.save('inverted_'+ sys.argv[1])

