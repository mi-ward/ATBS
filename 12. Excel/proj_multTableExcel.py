#! python3

# proj_multTableExcel.py - produces a multiplication table

import sys 
from openpyxl import Workbook
from openpyxl import styles

size = int(sys.argv[1])

wb = Workbook()

ws = wb.active
ws.title = ('Multiples of %s' % size)
font = styles.Font(bold=True)

# make header rows and bold them

for i in range(1, size + 1):
    ws.cell(row=1, column=(i+1), value=i).font = font
    ws.cell(row=(i+1), column=1, value=i).font = font

# compute values and store

for x in range(1, size + 1):
    for y in range(1, size + 1):
        product = ws.cell(row=1, column=(y+1)).value * ws.cell(row=(x+1), column=1).value
        ws.cell(row=(x+1), column=(y+1), value=product)

# save file
wb.save('multTable%s.xlsx' % str(size))
        
    









